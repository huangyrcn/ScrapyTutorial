from itemadapter import ItemAdapter

import sqlite3
from datetime import datetime
from openpyxl import Workbook
import lxml.html
from urllib.parse import urljoin
import re
import os

class NewsPipeline:
    def process_item(self, item, spider):
        """处理每个抓取的新闻项"""
        adapter = ItemAdapter(item)

        # 清洗数据
        # 1. 标题清洗：删除多余空格
        if adapter.get("title"):
            adapter["title"] = adapter["title"].strip()

        # 2. 发布日期清洗与格式化
        if adapter.get("publish_date"):
            adapter["publish_date"] = adapter["publish_date"].strip()
            # 可以添加日期格式化逻辑，例如将 "2025-03-18" 转为标准日期格式

        # 3. 作者/来源清洗
        if adapter.get("author"):
            adapter["author"] = adapter["author"].strip()

        # 添加爬取时间
        adapter["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return item


class SQLitePipeline:
    def __init__(self):
        # 数据库连接和游标
        self.conn = None
        self.cur = None

    def open_spider(self, spider):
        """当爬虫启动时创建数据库连接"""
        self.conn = sqlite3.connect("news.db")
        self.cur = self.conn.cursor()

        # 创建表（如果不存在）·
        self.cur.execute(
            """
            CREATE TABLE IF NOT EXISTS news (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT,
                publish_date TEXT,
                author TEXT,
                url TEXT UNIQUE,
                created_at TEXT
            )
        """
        )
        self.conn.commit()

    def close_spider(self, spider):
        """当爬虫关闭时关闭数据库连接"""
        self.conn.close()

    def process_item(self, item, spider):
        """将数据项插入数据库"""
        adapter = ItemAdapter(item)

        # 准备SQL语句和参数
        self.cur.execute(
            "INSERT OR IGNORE INTO news (title, publish_date, author, url, created_at) VALUES (?, ?, ?, ?, ?)",
            (
                adapter.get("title", ""),
                adapter.get("publish_date", ""),
                adapter.get("author", ""),
                adapter.get("url", ""),
                adapter.get("created_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ),
        )
        self.conn.commit()

        return item


class ExcelPipeline:
    def __init__(self):
        self.workbook = None
        self.sheet = None
        self.file_name = "news.xlsx"
        self.current_row = 1  # Start from row 2 (after header)
        print("ExcelPipeline init")

    def open_spider(self, spider):
        """当爬虫启动时创建Excel文件"""
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "新闻数据"

        # 设置表头
        headers = ["标题", "发布日期", "作者/来源", "URL", "爬取时间"]
        for col_num, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=col_num).value = header

    def close_spider(self, spider):
        self.workbook.save(self.file_name)

    def process_item(self, item, spider):
        """将数据项添加到Excel文件"""
        adapter = ItemAdapter(item)
        self.current_row += 1

        # 添加数据到对应的列
        self.sheet.cell(row=self.current_row, column=1).value = adapter.get("title", "")
        self.sheet.cell(row=self.current_row, column=2).value = adapter.get(
            "publish_date", ""
        )
        self.sheet.cell(row=self.current_row, column=3).value = adapter.get(
            "author", ""
        )
        self.sheet.cell(row=self.current_row, column=4).value = adapter.get("url", "")
        self.sheet.cell(row=self.current_row, column=5).value = adapter.get(
            "created_at", ""
        )

        return item



class HtmlSavePipeline:
    def __init__(self):
        # 创建保存目录
        self.output_dir = "html_files"
        os.makedirs(self.output_dir, exist_ok=True)
        
    def process_item(self, item, spider):
        """保存 HTML 并将相对路径转换为绝对路径"""
        adapter = ItemAdapter(item)
            
        response = adapter["response"]
        url = response.url
        
        try:
            # 解析 HTML
            doc = lxml.html.fromstring(response.text)
            
            # 转换 img 标签的 src
            for img in doc.xpath('//img'):
                src = img.get('src')
                if src and not src.startswith(('http://', 'https://', '//')):
                    img.set('src', urljoin(url, src))
            
            # 转换 a 标签的 href
            for a in doc.xpath('//a'):
                href = a.get('href')
                if href and not href.startswith(('http://', 'https://', '//', '#', 'javascript:')):
                    a.set('href', urljoin(url, href))
            
            # 转换 link 标签的 href
            for link in doc.xpath('//link'):
                href = link.get('href')
                if href and not href.startswith(('http://', 'https://', '//')):
                    link.set('href', urljoin(url, href))
            
            # 转换 script 标签的 src
            for script in doc.xpath('//script'):
                src = script.get('src')
                if src and not src.startswith(('http://', 'https://', '//')):
                    script.set('src', urljoin(url, src))
            
            # 获取修改后的 HTML
            html_content = lxml.html.tostring(doc, encoding='unicode', method='html')
            
            # 生成文件名 (使用标题或 URL 中的一部分)
            title = adapter.get('title', '')
            if not title:
                # 从 URL 创建文件名
                title = url.split('/')[-1].split('.')[0]
            
            # 清理文件名中的非法字符
            filename = re.sub(r'[\\/*?:"<>|]', "_", title)
            filename = filename[:100]  # 限制文件名长度
            
            # 保存 HTML 文件
            file_path = os.path.join(self.output_dir, f"{filename}.html")
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            # 将保存路径添加到 item 中
            adapter['html_saved_path'] = file_path
            
        except Exception as e:
            spider.logger.error(f"保存 HTML 时发生错误: {e}")
            
        return item